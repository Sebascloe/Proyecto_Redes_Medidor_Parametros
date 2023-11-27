import tkinter as tk
from tkinter import ttk
import ping3
import speedtest
import subprocess
from PIL import Image, ImageTk, ImageSequence
import pygame
import openpyxl
import threading
import re
import requests  # Necesitas instalar requests usando pip install requests




class App:
    def __init__(self, master):
        self.master = master
        master.title("Medición de velocidad de Internet")


         # Añadir una clave del canal ThingSpeak
        self.thingspeak_api_key = 'UDCP002E9BP59PNY'


        # Programar la primera medición después de 1 minuto (60,000 milisegundos)
        self.master.after(60000, self.start_measurement_periodic)


        # Añadir un ícono a la aplicación con Pillow
        icon_image = Image.open("icon.jpg")  # Asegúrate de que "icon.jpg" esté en la misma carpeta que tu script
        icon_image = ImageTk.PhotoImage(icon_image)
        master.iconphoto(False, icon_image)

        # Configurar el tema oscuro por defecto
        style = ttk.Style()
        style.theme_use('clam')  # 'clam' es un tema que proporciona un aspecto oscuro en sistemas Windows
        self.theme = 'dark'  # Mantener un seguimiento del tema actual
        
        # Crear un marco para el título
        title_frame = tk.Frame(master)
        title_frame.pack()




        # Agregar un emoji de "carita con dinero" a cada lado del título
        left_emoji_label = tk.Label(title_frame, text="Medidor", font=("Pacifico", 16))
        left_emoji_label.pack(pady=10)




        # Crear un marco para la imagen
        imagen_marco = tk.Frame(master)
        imagen_marco.pack(fill=tk.BOTH, expand=True)




        # Cargar y mostrar el GIF
        self.gif_frames = self.load_gif_frames('forest.gif')
        self.label = tk.Label(imagen_marco)
        self.label.pack(pady=10)




        # Iniciar la animación
        self.frame_number = 0
        self.animate()




        self.error_label = tk.Label(master, text="", fg="red")
        self.error_label.pack()




        # Crear un marco para la entrada de la dirección IP
        ip_frame = tk.Frame(master)
        ip_frame.pack()




        self.ip_label = tk.Label(ip_frame, text="Dirección IP:")
        self.ip_label.grid(row=0, column=0, padx=(0, 10))




        self.ip_entry = tk.Entry(ip_frame)
        self.ip_entry.grid(row=0, column=1)




        # Crear un marco para el label "Región" y el botón de región
        region_frame = tk.Frame(master)
        region_frame.pack()




        self.region_label = tk.Label(region_frame, text="Región:")
        self.region_label.grid(row=0, column=0, padx=(0, 10))




        self.region_var = tk.StringVar(value="América")
        self.region_selector = ttk.OptionMenu(region_frame, self.region_var, "Region", "America", "Asia", "Europa", style='AccentButton.TButton')
        self.region_selector.grid(row=0, column=1)




        # Crear un botón para medir la velocidad de Internet y guardar en Excel
        self.ping_button = ttk.Button(master, text="Medición", command=self.start_measurement, style='AccentButton.TButton')
        self.ping_button.pack()


        # Crear etiquetas para mostrar los resultados
        self.velocidad_label = tk.Label(master, text="Velocidad:")
        self.velocidad_label.pack(side=tk.LEFT, padx=(10, 10))




        self.latencia_label = tk.Label(master, text="Latencia:")
        self.latencia_label.pack(side=tk.LEFT, padx=(10, 10))




        self.perdida_label = tk.Label(master, text="Pérdida de Paquetes:")
        self.perdida_label.pack(side=tk.LEFT, padx=(10, 10))




        # Crear un botón para reproducir audio
        self.audio_button = ttk.Button(master, text="Reproducir audio", command=self.play_audio, style='AccentButton.TButton')
        self.audio_button.pack(pady=10, padx=10)

        # Crear un botón para cambiar entre temas oscuros y claros
        self.theme_button = ttk.Button(master, text="Cambiar Tema", command=self.toggle_theme)
        self.theme_button.pack(pady=10)


        # Variable para almacenar los resultados
        self.measurement_results = None
        
        # Configurar estilos
        #style = ttk.Style()
        #style.configure('AccentButton.TButton', foreground='#333333', background='#333333', font=('Segoe UI', 8))
        
        # Lista de nombres de archivos de audio
        self.audio_files = ["A.mp3","afton.mp3", "MrBeast.mp3","Frog.mp3","OV1.mp3", "OV4.mp3","CL.mp3","OS.mp3", "SH.mp3","TEN.mp3","Sans.mp3","Jungle.mp3"]
        self.current_audio_index = 0  # Índice actual en la lista de archivos de audio


    def toggle_theme(self):
        # Cambiar entre temas oscuros y claros
        if self.theme == 'dark':
            self.master.tk_setPalette(background='#ececec', foreground='black')
            style = ttk.Style()
            style.theme_use('default')  # Usar el tema predeterminado del sistema
            self.theme = 'light'
        else:
            self.master.tk_setPalette(background='#333333', foreground='#ffffff')
            style = ttk.Style()
            style.theme_use('clam')  # 'clam' es un tema oscuro en sistemas Windows
            self.theme = 'dark'
    
    

    def load_gif_frames(self, gif_path):
        gif = Image.open(gif_path)
        frames = [ImageTk.PhotoImage(frame) for frame in ImageSequence.Iterator(gif)]
        return frames




    def animate(self):
        frame = self.gif_frames[self.frame_number]
        self.label.configure(image=frame)




        # Avanza al siguiente frame
        self.frame_number += 1
        if self.frame_number == len(self.gif_frames):
            self.frame_number = 0




        # Programa el próximo frame
        self.master.after(50, self.animate)




    def play_audio(self):
        #pygame.mixer.init()
        #pygame.mixer.music.load('afton.mp3')
        #pygame.mixer.music.play()
        pygame.mixer.init()
        
        # Obtener el nombre del archivo de audio actual
        current_audio_file = self.audio_files[self.current_audio_index]
        
        pygame.mixer.music.load(current_audio_file)
        pygame.mixer.music.play()

        # Avanzar al siguiente archivo de audio en la lista
        self.current_audio_index += 1
        if self.current_audio_index == len(self.audio_files):
            self.current_audio_index = 0




       
   
    def start_measurement(self):
        # Validar la entrada antes de iniciar la medición
        ip = self.ip_entry.get()


        if not ip:
            self.error_label.config(text="Error: No se ha ingresado ninguna IP", fg="red")
            return
        
        # Utilizar expresiones regulares para validar la dirección IP
        if not self.is_valid_ip(ip):
            self.error_label.config(text="Error: Letras no permitidas", fg="red")
            return




        if not self.is_valid_ip_range(ip):
            self.error_label.config(text="Error: IP no correcta", fg="red")
            return




        # Mostrar el mensaje de "Procesando" en color verde
        self.error_label.config(text="Procesando", fg="green")




        # Iniciar la medición en un hilo
        measurement_thread = threading.Thread(target=self.perform_measurement)
        measurement_thread.start()


        # Programar la próxima medición después de 1 minuto (60,000 milisegundos)
        self.master.after(60000, self.start_measurement_periodic)

        
        
    def start_measurement_periodic(self):
        # Llamada a la función de medición
        self.start_measurement()


    def is_valid_ip(self, ip):
        # Utilizar expresiones regulares para validar la dirección IP
        ip_pattern = re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$')
        return bool(ip_pattern.match(ip))



    def is_valid_ip_range(self, ip):
        # Verificar si los números en la dirección IP están en el rango correcto
        return all(0 <= int(x) <= 255 for x in ip.split('.'))




    def perform_measurement(self):
        # Obtener la dirección IP y la región seleccionada
        ip = self.ip_entry.get()
        region = self.region_var.get()




        # Hacer ping a la dirección IP y medir la latencia y la pérdida de paquetes
        response_time = ping3.ping(ip)
        packet_loss = None
        with subprocess.Popen(['ping', '-n', '5', ip], stdout=subprocess.PIPE) as proc:
            for line in proc.stdout:
                line = line.decode('utf-8')
                if 'packet loss' in line:
                    packet_loss = float(line.split()[5].strip('%'))
                    break




        # Medir la velocidad de descarga y carga
        st = speedtest.Speedtest()
        download_speed = st.download() / 1000000
        upload_speed = st.upload() / 1000000




        # Guardar los resultados en un archivo Excel
        self.save_to_excel(ip, region, response_time, packet_loss, download_speed, upload_speed)




        # Actualizar la interfaz gráfica con los resultados
        self.measurement_results = (response_time, packet_loss, download_speed, upload_speed)
        self.master.after(0, self.update_ui)


        # Enviar datos a ThingSpeak
        self.send_to_thingspeak(download_speed, upload_speed, response_time, packet_loss)


    def send_to_thingspeak(self, download_speed, upload_speed, response_time, packet_loss):
        # URL del canal ThingSpeak
        thingspeak_url = f'https://api.thingspeak.com/update?api_key={self.thingspeak_api_key}'


        # Construir el cuerpo de la solicitud con los datos de medición
        data = {
            'field1': download_speed,
            'field2': upload_speed,
            'field3': response_time,
            'field4': packet_loss,
        }


        # Enviar la solicitud HTTP POST a ThingSpeak
        try:
            response = requests.post(thingspeak_url, data=data)
            if response.status_code == 200:
                print("Datos enviados a ThingSpeak con éxito.")
            else:
                print(f"Error al enviar datos a ThingSpeak. Código de estado: {response.status_code}")
        except Exception as e:
            print(f"Error en la solicitud HTTP: {e}")




    def update_ui(self):
        # Mostrar los resultados debajo de cada etiqueta
        response_time, packet_loss, download_speed, upload_speed = self.measurement_results
        self.velocidad_label.config(
            text=f"Velocidad: Descarga: {download_speed:.2f} Mbps, Carga: {upload_speed:.2f} Mbps")
        self.latencia_label.config(text=f"Latencia: {response_time} ms")
        self.perdida_label.config(text=f"Pérdida de Paquetes: {packet_loss}%")
        self.error_label.config(text="")




    def save_to_excel(self, ip, region, response_time, packet_loss, download_speed, upload_speed):
        # Cargar el archivo Excel existente o crear uno nuevo
        try:
            wb = openpyxl.load_workbook("resultados18.xlsx")
            ws = wb.active
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            ws = wb.active
            # Escribir encabezados solo si el archivo es nuevo
            ws.append(["Dirección IP", "Región", "Latencia", "Pérdida de Paquetes", "Velocidad de Descarga", "Velocidad de Carga"])




        # Escribir datos en la siguiente fila disponible
        last_row = ws.max_row + 1
        ws.cell(row=last_row, column=1, value=ip)
        ws.cell(row=last_row, column=2, value=region)
        ws.cell(row=last_row, column=3, value=response_time or 0)
        ws.cell(row=last_row, column=4, value=packet_loss or 0)
        ws.cell(row=last_row, column=5, value=download_speed or 0)
        ws.cell(row=last_row, column=6, value=upload_speed or 0)




        # Guardar el archivo Excel
        wb.save("resultados20.xlsx")




        # Actualizar la interfaz gráfica con los resultados
        self.measurement_results = (response_time, packet_loss, download_speed, upload_speed)
        self.master.after(0, self.update_ui)




# Resto del código sin cambios




root = tk.Tk()
app = App(root)
root.mainloop()
